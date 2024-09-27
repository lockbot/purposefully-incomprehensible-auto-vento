import os
import sqlite3
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook

def get_db_path():
    """Get the path to the SQLite database."""
    app_data = os.getenv('APPDATA')
    if not app_data:
        raise Exception("APPDATA environment variable not found.")
    db_path = os.path.join(app_data, 'HermetoPascoal', 'db', 'vento-1.3.2.db')
    if not os.path.exists(db_path):
        raise Exception(f"Database file not found at {db_path}")
    return db_path


def rescue_and_export_breaths():
    """
    Rescues all breaths and breath_gas entries by setting deleted_at = NULL where it's not NULL.
    Sets breath_confirm and alert_start for each breath appropriately.
    Exports the data to an Excel file, adding a new worksheet with the current timestamp.
    Ensures that the most recent worksheet is the first one (leftmost).
    """
    db_path = get_db_path()
    conn = sqlite3.connect(db_path)
    try:
        cursor = conn.cursor()

        # Set deleted_at = NULL for all breaths and breath_gas entries
        cursor.execute("""
            UPDATE breaths
            SET deleted_at = NULL
            WHERE deleted_at IS NOT NULL
        """)
        cursor.execute("""
            UPDATE breath_gas
            SET deleted_at = NULL
            WHERE deleted_at IS NOT NULL
        """)
        conn.commit()

        # Get all breaths ordered by exam_id and id
        cursor.execute("""
            SELECT id, exam_id, date_time_finish
            FROM breaths
            ORDER BY exam_id, id
        """)
        breaths = cursor.fetchall()

        # Process breaths grouped by exam_id
        from collections import defaultdict

        breaths_by_exam = defaultdict(list)
        for breath in breaths:
            breath_id, exam_id, date_time_finish = breath
            breaths_by_exam[exam_id].append((breath_id, date_time_finish))

        for exam_id, breaths_list in breaths_by_exam.items():
            num_breaths = len(breaths_list)
            for idx, (breath_id, date_time_finish) in enumerate(breaths_list):
                if date_time_finish is None:
                    raise Exception(f"date_time_finish is NULL for breath id {breath_id}")

                try:
                    # Parse date_time_finish
                    date_time_finish_dt = datetime.fromisoformat(date_time_finish.replace(' ', 'T'))
                except Exception as e:
                    raise Exception(f"Invalid date_time_finish format for breath id {breath_id}: {date_time_finish}") from e

                # Set breath_confirm to date_time_finish + 1 second
                breath_confirm_dt = date_time_finish_dt + timedelta(seconds=1)
                breath_confirm_str = breath_confirm_dt.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]

                # For the 100th breath, set alert_start to NULL
                if num_breaths == 100 and idx == num_breaths - 1:
                    alert_start_str = None
                else:
                    # Set alert_start to breath_confirm + 1 second
                    alert_start_dt = breath_confirm_dt + timedelta(seconds=1)
                    alert_start_str = alert_start_dt.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]

                # Update the breath
                cursor.execute("""
                    UPDATE breaths
                    SET breath_confirm = ?,
                        alert_start = ?
                    WHERE id = ?
                """, (breath_confirm_str, alert_start_str, breath_id))
        conn.commit()

        # Export data to Excel
        export_to_excel(conn)

    finally:
        conn.close()


def get_gas_name(gas_id):
    """Convert gas_id to gas name."""
    if gas_id == 1:
        return "H2"
    elif gas_id == 2:
        return "CH4"
    else:
        return "unknown"


def export_to_excel(conn):
    """
    Exports the breaths data to an Excel file, adding a new worksheet with the current timestamp.
    Ensures that the most recent worksheet is the first one (leftmost).
    """
    cursor = conn.cursor()
    cursor.execute("""
        SELECT b.exam_id, b.id as breath_id, bg.id as breath_gas_id,
               bg.created_at, bg.gas, bg.ppm
        FROM breaths b
        JOIN breath_gas bg ON b.id = bg.breath_id
        WHERE b.deleted_at IS NULL
        ORDER BY b.exam_id DESC, b.id, bg.id
    """)
    breaths = cursor.fetchall()

    if not breaths:
        print("No breath data found where deleted_at is NULL.")
        return

    # Define the timestamp for the worksheet name
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    sheet_name = timestamp

    # Define the output Excel filename
    output_filename = "breaths_data.xlsx"
    # Save the file on the Desktop
    output_filename = os.path.join(os.path.expanduser("~"), "Desktop", output_filename)

    # Check if the Excel file exists
    if os.path.exists(output_filename):
        # Load the existing workbook
        workbook = load_workbook(output_filename)
    else:
        # Create a new workbook
        workbook = Workbook()
        # Remove the default sheet
        default_sheet = workbook.active
        workbook.remove(default_sheet)

    # Add a new worksheet at the first position
    worksheet = workbook.create_sheet(title=sheet_name, index=0)

    # Write the headers
    headers = ['Exam ID', 'Breath ID', 'BreathGas ID', 'Created At', 'Gas', 'PPM']
    worksheet.append(headers)

    # Write the data
    for row in breaths:
        exam_id, breath_id, breath_gas_id, created_at, gas_id, ppm = row
        gas_name = get_gas_name(gas_id)
        worksheet.append([exam_id, breath_id, breath_gas_id, created_at, gas_name, ppm])

    # Save the workbook
    workbook.save(output_filename)
    print(f"Data exported successfully to {output_filename}")
